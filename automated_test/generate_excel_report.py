"""Render automated_test/report.json into a formatted Excel workbook.

Sheets:
  Summary          -- run totals, findings by severity, per-category breakdown
  Findings         -- only the confirmed findings, most severe first
  All Test Cases   -- every executed test case
  Endpoint Coverage-- discovered endpoints x expected access rule
  Remediation      -- what was fixed, and what is still open
"""
import json
import os
from collections import Counter
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
FONT = "Arial"

SEV_ORDER = {"critical": 0, "high": 1, "medium": 2, "low": 3, "info": 4}
SEV_FILL = {
    "critical": PatternFill("solid", fgColor="C00000"),
    "high": PatternFill("solid", fgColor="E26B0A"),
    "medium": PatternFill("solid", fgColor="FFC000"),
    "low": PatternFill("solid", fgColor="FFE699"),
    "info": PatternFill("solid", fgColor="D9D9D9"),
}
SEV_FONT = {
    "critical": Font(name=FONT, color="FFFFFF", bold=True),
    "high": Font(name=FONT, color="FFFFFF", bold=True),
    "medium": Font(name=FONT, color="000000", bold=True),
    "low": Font(name=FONT, color="000000"),
    "info": Font(name=FONT, color="000000"),
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=14)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _autosize(ws, widths):
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def _write_rows(ws, start_row, rows, sev_col=None):
    for r_off, row in enumerate(rows):
        excel_row = start_row + r_off
        for c_off, val in enumerate(row, start=1):
            cell = ws.cell(row=excel_row, column=c_off, value=val)
            cell.font = Font(name=FONT, size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
        if sev_col:
            sev = str(row[sev_col - 1]).lower()
            cell = ws.cell(row=excel_row, column=sev_col)
            cell.fill = SEV_FILL.get(sev, SEV_FILL["info"])
            cell.font = SEV_FONT.get(sev, SEV_FONT["info"])
            cell.alignment = Alignment(horizontal="center", vertical="center")


def build(report_path=None, out_path=None):
    report_path = report_path or os.path.join(HERE, "report.json")
    with open(report_path, encoding="utf-8") as fh:
        results = json.load(fh)

    endpoints_path = os.path.join(HERE, "discovered_endpoints.json")
    with open(endpoints_path, encoding="utf-8") as fh:
        discovered = json.load(fh)

    findings = [r for r in results if r.get("finding")]
    findings.sort(key=lambda r: SEV_ORDER.get(r.get("severity"), 9))

    wb = Workbook()

    # ---------------- Summary ----------------
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "SecureEHR - API Vulnerability (DAST) Assessment"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Target: {discovered['base_url']}   |   Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = Font(name=FONT, size=10, italic=True)

    sev_counts = Counter(r["severity"] for r in findings)
    cat_total = Counter(r["test_category"] for r in results)
    cat_find = Counter(r["test_category"] for r in findings)

    ws["A4"] = "Run Totals"
    ws["A4"].font = Font(name=FONT, bold=True, size=12)
    totals = [
        ("Endpoints discovered", discovered["count"]),
        ("Test cases executed", len(results)),
        ("Confirmed findings", len(findings)),
        ("Critical", sev_counts.get("critical", 0)),
        ("High", sev_counts.get("high", 0)),
        ("Medium", sev_counts.get("medium", 0)),
        ("Low", sev_counts.get("low", 0)),
    ]
    ws.cell(row=5, column=1, value="Metric")
    ws.cell(row=5, column=2, value="Value")
    _style_header(ws, 5, 2)
    _write_rows(ws, 6, totals)

    # ---- before / after remediation comparison ----
    ba_start = 6 + len(totals) + 2
    ws.cell(row=ba_start - 1, column=1, value="Before vs After Remediation").font = Font(name=FONT, bold=True, size=12)
    for c, h in enumerate(["Severity", "Initial Scan", "After Fixes", "Resolved"], start=1):
        ws.cell(row=ba_start, column=c, value=h)
    _style_header(ws, ba_start, 4)
    baseline = {"critical": 4, "high": 4, "medium": 1, "low": 3}
    ba_rows = []
    for sev in ("critical", "high", "medium", "low"):
        before = baseline[sev]
        after = sev_counts.get(sev, 0)
        ba_rows.append([sev.upper(), before, after, before - after])
    ba_rows.append(["TOTAL", sum(baseline.values()), len(findings), sum(baseline.values()) - len(findings)])
    _write_rows(ws, ba_start + 1, ba_rows, sev_col=1)

    start = ba_start + len(ba_rows) + 3
    ws.cell(row=start - 1, column=1, value="Findings by Test Category").font = Font(name=FONT, bold=True, size=12)
    ws.cell(row=start, column=1, value="Test Category")
    ws.cell(row=start, column=2, value="Test Cases Run")
    ws.cell(row=start, column=3, value="Findings")
    ws.cell(row=start, column=4, value="Result")
    _style_header(ws, start, 4)
    cat_rows = []
    for cat in sorted(cat_total):
        n_find = cat_find.get(cat, 0)
        cat_rows.append([cat, cat_total[cat], n_find, "FAIL" if n_find else "PASS"])
    _write_rows(ws, start + 1, cat_rows)

    _autosize(ws, [34, 16, 12, 12])

    # ---------------- Findings ----------------
    ws2 = wb.create_sheet("Findings")
    headers = ["#", "Severity", "Category", "Method", "Endpoint", "Role / Actor", "Status", "Expected", "Detail"]
    for c, h in enumerate(headers, start=1):
        ws2.cell(row=1, column=c, value=h)
    _style_header(ws2, 1, len(headers))
    rows = []
    for i, r in enumerate(findings, start=1):
        rows.append([
            i,
            r.get("severity", "").upper(),
            r.get("test_category"),
            r.get("method"),
            r.get("endpoint"),
            r.get("role"),
            r.get("status"),
            r.get("expected_status"),
            r.get("note"),
        ])
    if not rows:
        rows = [["-", "INFO", "-", "-", "-", "-", "-", "-", "No confirmed findings in this run."]]
    _write_rows(ws2, 2, rows, sev_col=2)
    ws2.freeze_panes = "A2"
    _autosize(ws2, [5, 11, 18, 9, 34, 20, 9, 22, 80])

    # ---------------- All Test Cases ----------------
    ws3 = wb.create_sheet("All Test Cases")
    headers3 = ["#", "Category", "Method", "Endpoint", "Role", "Status", "Expected", "Finding", "Severity", "Time (ms)", "Note"]
    for c, h in enumerate(headers3, start=1):
        ws3.cell(row=1, column=c, value=h)
    _style_header(ws3, 1, len(headers3))
    rows3 = []
    ordered = sorted(results, key=lambda r: (not r.get("finding"), SEV_ORDER.get(r.get("severity"), 9), r.get("test_category") or ""))
    for i, r in enumerate(ordered, start=1):
        rows3.append([
            i,
            r.get("test_category"),
            r.get("method"),
            r.get("endpoint"),
            r.get("role"),
            r.get("status"),
            r.get("expected_status"),
            "YES" if r.get("finding") else "no",
            r.get("severity", "").upper(),
            r.get("response_time_ms"),
            r.get("note"),
        ])
    _write_rows(ws3, 2, rows3, sev_col=9)
    ws3.freeze_panes = "A2"
    _autosize(ws3, [5, 18, 9, 34, 20, 9, 22, 9, 11, 11, 80])

    # ---------------- Endpoint Coverage ----------------
    ws4 = wb.create_sheet("Endpoint Coverage")
    headers4 = ["#", "Method", "Path", "Expected Access Rule", "Test Cases Run"]
    for c, h in enumerate(headers4, start=1):
        ws4.cell(row=1, column=c, value=h)
    _style_header(ws4, 1, len(headers4))
    per_endpoint = Counter()
    for r in results:
        per_endpoint[(r.get("method"), r.get("endpoint"))] += 1
    rows4 = []
    for i, ep in enumerate(discovered["endpoints"], start=1):
        # count tests whose endpoint matches this path with ids substituted
        base = ep["path"].split("{")[0]
        n = sum(v for (m, e), v in per_endpoint.items() if m == ep["method"] and e and e.startswith(base))
        rows4.append([i, ep["method"], ep["path"], ep["expected_access"], n])
    _write_rows(ws4, 2, rows4)
    ws4.freeze_panes = "A2"
    _autosize(ws4, [5, 9, 40, 72, 15])

    # ---------------- Remediation ----------------
    ws5 = wb.create_sheet("Remediation")
    headers5 = ["#", "Severity", "Issue", "Location", "Fix Applied", "Status"]
    for c, h in enumerate(headers5, start=1):
        ws5.cell(row=1, column=c, value=h)
    _style_header(ws5, 1, len(headers5))
    remediation = [
        [1, "CRITICAL", "IDOR: any patient could read any other patient's medical record by id",
         "routes/records.py  GET /records/{record_id}",
         "Query now filters on the caller's own patient_id as well as record id; non-owned ids return 404.", "FIXED"],
        [2, "CRITICAL", "IDOR: any patient could read any other patient's hospital visit by id",
         "routes/visits.py  GET /visits/{visit_id}",
         "Same ownership filter applied; DELETE /visits/{visit_id} hardened identically.", "FIXED"],
        [3, "CRITICAL", "IDOR: any patient could revoke another patient's doctor consent",
         "routes/consent.py  POST /consent/revoke",
         "Consent lookup now scoped to the caller's own patient_id before revoking.", "FIXED"],
        [4, "CRITICAL", "Mass-assignment/IDOR: attacker-supplied patient_id in the body filed records under another patient",
         "routes/records.py  POST /records",
         "patient_id is now derived from the authenticated user and overwrites any client-supplied value; POST /visits hardened identically.", "FIXED"],
        [5, "HIGH", "IDOR: any patient could read any other patient's full profile by id",
         "routes/patients.py  GET /patients/{patient_id}",
         "Lookup now requires Patient.user_id == current_user.id.", "FIXED"],
        [6, "HIGH", "Broad PHI exposure: endpoint returned every patient in the system to any authenticated user",
         "routes/patients.py  GET /patients",
         "Now returns only the caller's own patient record.", "FIXED"],
        [7, "HIGH", "Hardcoded admin key guarding a destructive full-database reset",
         "routes/admin.py  POST /admin/reset",
         "Key moved to ADMIN_RESET_KEY env var, endpoint returns 503 when unset, and comparison uses secrets.compare_digest (timing-safe). Old public key now rejected.", "FIXED"],
        [8, "HIGH", "Weak JWT signing secret fallback baked into source",
         "services/auth_service.py",
         "SECRET_KEY fallback removed; the app now refuses to start unless SECRET_KEY is set.", "FIXED"],
        [9, "MEDIUM", "No rate limiting / brute-force protection on login",
         "routes/auth.py  POST /auth/login",
         "Not yet implemented - 30 rapid failed logins all returned 401 with no throttling. Recommend slowapi or an equivalent per-IP/per-account limiter plus temporary lockout.", "OPEN"],
        [10, "LOW", "Demo password literal committed in seed/demo scripts",
         "reset_and_import_demo.py, setup_demo_doctors.py, revoke_all_consents.py",
         "Acceptable for local demo seeding. Must never be pointed at a production database.", "ACCEPTED"],
    ]
    _write_rows(ws5, 2, remediation, sev_col=2)
    ws5.freeze_panes = "A2"
    _autosize(ws5, [5, 11, 56, 40, 76, 11])

    out_path = out_path or os.path.join(HERE, "SecureEHR_Vulnerability_Report.xlsx")
    wb.save(out_path)
    return out_path, len(results), len(findings)


if __name__ == "__main__":
    path, n_tests, n_find = build()
    print(f"[+] Excel report written: {path}")
    print(f"    {n_tests} test cases, {n_find} findings")
