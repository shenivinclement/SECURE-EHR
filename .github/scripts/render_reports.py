"""Render the SecureEHR test-report spreadsheets into the GitHub Actions
run summary, so results are visible in the Actions tab without downloading
anything. The same .xlsx files are also uploaded as build artifacts.

Writes GitHub-flavoured Markdown to $GITHUB_STEP_SUMMARY (falls back to
stdout when run locally).
"""
import os
import sys
from datetime import datetime, timezone

from openpyxl import load_workbook

# (path, friendly title, [sheets rendered in full])
# Sheets not listed are rendered truncated to MAX_ROWS.
REPORTS = [
    ("E2E_Test_Report_SecureEHR.xlsx",
     "Web E2E — Selenium (300 cases)",
     ["Summary", "Failed Tests"]),
    ("Appium_Mobile_E2E_Test_Report_SecureEHR.xlsx",
     "Android Mobile E2E — Appium (310 cases)",
     ["Executive Dashboard", "Deployable Readiness Summary"]),
    ("Load_Test_Report_SecureEHR.xlsx",
     "API Load & Baseline Performance (100 VUs)",
     ["Executive Load Dashboard", "System Load Readiness Summary"]),
    ("automated_test/SecureEHR_Vulnerability_Report.xlsx",
     "API Security — DAST / Vulnerability Assessment",
     ["Summary", "Findings", "Remediation"]),
]

MAX_ROWS = 25       # truncation cap for non-highlighted sheets
MAX_COLS = 12
MAX_CELL = 160      # characters per cell before ellipsis


def esc(value):
    """Make a cell value safe for a Markdown table cell."""
    if value is None:
        return ""
    text = str(value).replace("\r", " ").replace("\n", " ").strip()
    text = text.replace("|", "\\|")
    if len(text) > MAX_CELL:
        text = text[: MAX_CELL - 1] + "…"
    return text


def sheet_to_markdown(ws, full=False):
    """Render a worksheet as a Markdown table. Returns (lines, truncated)."""
    rows = []
    for row in ws.iter_rows(values_only=True):
        cells = [esc(v) for v in row[:MAX_COLS]]
        if any(c for c in cells):
            rows.append(cells)

    if not rows:
        return ["_(empty sheet)_", ""], False

    limit = len(rows) if full else min(len(rows), MAX_ROWS)
    truncated = limit < len(rows)
    shown = rows[:limit]

    width = max(len(r) for r in shown)
    shown = [r + [""] * (width - len(r)) for r in shown]

    header, *body = shown
    # a title-style banner row (single populated cell) reads badly as a header
    if sum(1 for c in header if c) == 1 and body:
        lines = [f"**{header[0]}**", ""]
        header, body = body[0], body[1:]
        width = max(len(header), max((len(r) for r in body), default=0))
        header = header + [""] * (width - len(header))
        body = [r + [""] * (width - len(r)) for r in body]
    else:
        lines = []

    header = [h if h else f"Col{i+1}" for i, h in enumerate(header)]
    lines.append("| " + " | ".join(header) + " |")
    lines.append("|" + "|".join([" --- "] * len(header)) + "|")
    for r in body:
        lines.append("| " + " | ".join(r[: len(header)]) + " |")
    lines.append("")
    return lines, truncated


def render_report(path, title, highlight):
    lines = []
    if not os.path.exists(path):
        lines.append(f"## {title}")
        lines.append("")
        lines.append(f"> **Report not found** — `{path}` was not present in this checkout.")
        lines.append("")
        return lines, False

    wb = load_workbook(path, data_only=True)
    size_kb = os.path.getsize(path) / 1024

    lines.append(f"## {title}")
    lines.append("")
    lines.append(f"`{path}` · {len(wb.worksheets)} sheets · {size_kb:,.0f} KB")
    lines.append("")

    for ws in wb.worksheets:
        full = ws.title in highlight
        body, truncated = sheet_to_markdown(ws, full=full)
        if full:
            lines.append(f"### {ws.title}")
            lines.append("")
            lines.extend(body)
        else:
            # keep the big raw-data sheets collapsed so the page stays readable
            lines.append("<details>")
            lines.append(f"<summary><b>{ws.title}</b> — {ws.max_row} rows "
                         f"(showing first {min(MAX_ROWS, ws.max_row)})</summary>")
            lines.append("")
            lines.extend(body)
            if truncated:
                lines.append(f"_…truncated. Download the artifact for all {ws.max_row} rows._")
                lines.append("")
            lines.append("</details>")
            lines.append("")

    return lines, True


def main():
    out = []
    out.append("# SecureEHR — Automated Test Report Dashboard")
    out.append("")
    out.append(f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}")
    out.append("")
    out.append("> Every spreadsheet below is also published to the **Artifacts** "
               "section of this workflow run as a downloadable `.xlsx`.")
    out.append("")

    found, missing = [], []
    bodies = []
    for path, title, highlight in REPORTS:
        lines, ok = render_report(path, title, highlight)
        bodies.extend(lines)
        bodies.append("---")
        bodies.append("")
        (found if ok else missing).append(os.path.basename(path))

    out.append("| Report | Status |")
    out.append("| --- | --- |")
    for path, title, _ in REPORTS:
        status = "✅ published" if os.path.exists(path) else "⚠️ not found"
        out.append(f"| {title} | {status} |")
    out.append("")
    out.append("---")
    out.append("")
    out.extend(bodies)

    text = "\n".join(out)

    summary_path = os.environ.get("GITHUB_STEP_SUMMARY")
    if summary_path:
        with open(summary_path, "a", encoding="utf-8") as fh:
            fh.write(text)
        print(f"[+] Wrote {len(text):,} chars to the workflow summary.")
    else:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        print(text)

    print(f"[+] Rendered {len(found)} report(s): {', '.join(found) or 'none'}")
    if missing:
        print(f"[!] Missing {len(missing)} report(s): {', '.join(missing)}")


if __name__ == "__main__":
    main()
