import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import datetime

def build_excel_report(summary_data, passed_tests, failed_tests, execution_logs, test_details, output_path):
    """
    Generates a beautifully styled Excel E2E Test Report matching the reference format.
    Sheets:
    1. Summary
    2. Passed Tests
    3. Failed Tests
    4. Execution Log
    5. Test Details
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Styles
    font_family = "Segoe UI"
    
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    title_font = Font(name=font_family, size=14, bold=True, color="1E293B")
    bold_font = Font(name=font_family, size=10, bold=True, color="0F172A")
    normal_font = Font(name=font_family, size=10, color="334155")
    
    pass_font = Font(name=font_family, size=10, bold=True, color="15803D")
    fail_font = Font(name=font_family, size=10, bold=True, color="B91C1C")
    
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Navy Blue
    pass_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # Soft Green
    fail_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Soft Red
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    thin_border_side = Side(border_style="thin", color="CBD5E1")
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")

    # -------------------------------------------------------------
    # 1. SUMMARY SHEET
    # -------------------------------------------------------------
    ws_summary = wb.create_sheet(title="Summary")
    ws_summary.views.sheetView[0].showGridLines = True

    summary_headers = ["Test Suite", "Total Tests", "Passed", "Failed", "Pass Rate %", "Duration (sec)", "Start Time", "End Time"]
    ws_summary.append(summary_headers)
    
    summary_row = [
        summary_data.get("suite_name", "SECURE EHR Web App — Full E2E Workflow"),
        summary_data.get("total_tests", 300),
        summary_data.get("passed", 0),
        summary_data.get("failed", 0),
        summary_data.get("pass_rate", 0.0),
        summary_data.get("duration", 0.0),
        summary_data.get("start_time", ""),
        summary_data.get("end_time", "")
    ]
    ws_summary.append(summary_row)

    # Style Header
    for col_idx in range(1, len(summary_headers) + 1):
        cell = ws_summary.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = cell_border
        ws_summary.row_dimensions[1].height = 28

    # Style Data Row
    ws_summary.row_dimensions[2].height = 24
    for col_idx in range(1, len(summary_row) + 1):
        cell = ws_summary.cell(row=2, column=col_idx)
        cell.font = bold_font if col_idx in [2, 3, 4, 5] else normal_font
        cell.border = cell_border
        cell.alignment = align_center if col_idx in [2, 3, 4, 5, 6] else align_left
        if col_idx == 3:
            cell.font = pass_font
        elif col_idx == 4 and summary_data.get("failed", 0) > 0:
            cell.font = fail_font

    # -------------------------------------------------------------
    # 2. PASSED TESTS SHEET
    # -------------------------------------------------------------
    ws_passed = wb.create_sheet(title="Passed Tests")
    ws_passed.views.sheetView[0].showGridLines = True
    passed_headers = ["No.", "Category", "Test Name", "Time (sec)", "Status"]
    ws_passed.append(passed_headers)

    for col_idx in range(1, len(passed_headers) + 1):
        cell = ws_passed.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = cell_border
    ws_passed.row_dimensions[1].height = 26

    for idx, item in enumerate(passed_tests, start=1):
        row_num = idx + 1
        ws_passed.append([idx, item["category"], item["test_name"], item["duration"], item["status"]])
        ws_passed.row_dimensions[row_num].height = 20
        fill = zebra_fill if idx % 2 == 0 else PatternFill(fill_type=None)
        
        for c_idx in range(1, 6):
            cell = ws_passed.cell(row=row_num, column=c_idx)
            cell.font = normal_font
            cell.fill = fill
            cell.border = cell_border
            if c_idx in [1, 4]:
                cell.alignment = align_center
            elif c_idx == 5:
                cell.alignment = align_center
                cell.font = pass_font
                cell.fill = pass_fill
            else:
                cell.alignment = align_left

    # -------------------------------------------------------------
    # 3. FAILED TESTS SHEET
    # -------------------------------------------------------------
    ws_failed = wb.create_sheet(title="Failed Tests")
    ws_failed.views.sheetView[0].showGridLines = True
    failed_headers = ["No.", "Category", "Test Name", "Error", "Status", "Timestamp"]
    ws_failed.append(failed_headers)

    for col_idx in range(1, len(failed_headers) + 1):
        cell = ws_failed.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = PatternFill(start_color="991B1B", end_color="991B1B", fill_type="solid") # Dark Crimson Red
        cell.alignment = align_center
        cell.border = cell_border
    ws_failed.row_dimensions[1].height = 26

    for idx, item in enumerate(failed_tests, start=1):
        row_num = idx + 1
        ws_failed.append([idx, item["category"], item["test_name"], item["error"], item["status"], item["timestamp"]])
        ws_failed.row_dimensions[row_num].height = 24
        
        for c_idx in range(1, 7):
            cell = ws_failed.cell(row=row_num, column=c_idx)
            cell.font = normal_font
            cell.border = cell_border
            if c_idx in [1, 6]:
                cell.alignment = align_center
            elif c_idx == 5:
                cell.alignment = align_center
                cell.font = fail_font
                cell.fill = fail_fill
            else:
                cell.alignment = align_left

    # -------------------------------------------------------------
    # 4. EXECUTION LOG SHEET
    # -------------------------------------------------------------
    ws_log = wb.create_sheet(title="Execution Log")
    ws_log.views.sheetView[0].showGridLines = True
    log_headers = ["Timestamp", "Level", "Message"]
    ws_log.append(log_headers)

    for col_idx in range(1, len(log_headers) + 1):
        cell = ws_log.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = cell_border
    ws_log.row_dimensions[1].height = 26

    for idx, item in enumerate(execution_logs, start=1):
        row_num = idx + 1
        ws_log.append([item["timestamp"], item["level"], item["message"]])
        ws_log.row_dimensions[row_num].height = 19
        fill = zebra_fill if idx % 2 == 0 else PatternFill(fill_type=None)

        for c_idx in range(1, 4):
            cell = ws_log.cell(row=row_num, column=c_idx)
            cell.font = normal_font
            cell.fill = fill
            cell.border = cell_border
            if c_idx == 1:
                cell.alignment = align_center
            elif c_idx == 2:
                cell.alignment = align_center
                cell.font = bold_font
            else:
                cell.alignment = align_left

    # -------------------------------------------------------------
    # 5. TEST DETAILS SHEET
    # -------------------------------------------------------------
    ws_details = wb.create_sheet(title="Test Details")
    ws_details.views.sheetView[0].showGridLines = True
    details_headers = ["No.", "Category", "Test Name", "Status", "Error Details"]
    ws_details.append(details_headers)

    for col_idx in range(1, len(details_headers) + 1):
        cell = ws_details.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = cell_border
    ws_details.row_dimensions[1].height = 26

    for idx, item in enumerate(test_details, start=1):
        row_num = idx + 1
        ws_details.append([idx, item["category"], item["test_name"], item["status"], item["error_details"]])
        ws_details.row_dimensions[row_num].height = 20
        fill = zebra_fill if idx % 2 == 0 else PatternFill(fill_type=None)

        for c_idx in range(1, 6):
            cell = ws_details.cell(row=row_num, column=c_idx)
            cell.font = normal_font
            cell.fill = fill
            cell.border = cell_border
            if c_idx in [1]:
                cell.alignment = align_center
            elif c_idx == 4:
                cell.alignment = align_center
                if item["status"] == "PASSED":
                    cell.font = pass_font
                    cell.fill = pass_fill
                else:
                    cell.font = fail_font
                    cell.fill = fail_fill
            else:
                cell.alignment = align_left

    # Auto-adjust column widths across all sheets
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                # If multi-line, take longest line length
                lines = val_str.split('\n')
                line_len = max(len(l) for l in lines) if lines else 0
                max_len = max(max_len, line_len)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 70)

    # Save workbook
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"[+] Excel report generated successfully: {output_path}")

